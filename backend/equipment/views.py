import pandas as pd
import numpy as np
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import Dataset, EquipmentAlert, MaintenanceSchedule, EquipmentParameter, EmailReportSchedule, EquipmentRanking
from reportlab.lib.pagesizes import A4, letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.http import HttpResponse
from io import BytesIO
from datetime import datetime, timedelta
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from sklearn.linear_model import LinearRegression

# NO AUTH - All endpoints open

@api_view(['POST'])
def upload_csv(request):
    file = request.FILES.get('file')

    if not file:
        return Response({"error": "No file uploaded"}, status=400)

    try:
        df = pd.read_csv(file)
        
        required_columns = ['Equipment Name', 'Type', 'Flowrate', 'Pressure', 'Temperature']
        missing = [col for col in required_columns if col not in df.columns]
        if missing:
            return Response({"error": f"Missing columns: {', '.join(missing)}"}, status=400)

        summary = {
            "total_records": len(df),
            "avg_flowrate": float(df["Flowrate"].mean()),
            "avg_pressure": float(df["Pressure"].mean()),
            "avg_temperature": float(df["Temperature"].mean()),
            "type_distribution": df["Type"].value_counts().to_dict()
        }

        from django.utils import timezone
        dataset = Dataset.objects.create(
            total_records=summary["total_records"],
            avg_flowrate=summary["avg_flowrate"],
            avg_pressure=summary["avg_pressure"],
            avg_temperature=summary["avg_temperature"],
            file_name=file.name
        )

        for _, row in df.iterrows():
            flowrate = float(row['Flowrate'])
            pressure = float(row['Pressure'])
            temperature = float(row['Temperature'])
            
            flowrate_score = 100 if 100 <= flowrate <= 130 else 80 if 80 <= flowrate <= 150 else 60
            pressure_score = 100 if 4 <= pressure <= 8 else 80 if 3 <= pressure <= 9 else 60
            temp_score = 100 if 100 <= temperature <= 135 else 80 if 90 <= temperature <= 150 else 60
            health_score = (flowrate_score + pressure_score + temp_score) / 3
            
            EquipmentParameter.objects.create(
                dataset=dataset,
                equipment_name=row['Equipment Name'],
                equipment_type=row['Type'],
                flowrate=flowrate,
                pressure=pressure,
                temperature=temperature,
                health_score=health_score,
                efficiency_index=health_score
            )
            
            if flowrate > 150 or flowrate < 50:
                EquipmentAlert.objects.create(
                    equipment_name=row['Equipment Name'],
                    alert_type='CRITICAL',
                    parameter='Flowrate',
                    value=flowrate,
                    threshold=150 if flowrate > 150 else 50,
                    message=f"Flowrate {'critically high' if flowrate > 150 else 'critically low'}: {flowrate:.2f} L/min",
                    recommendation="Immediate inspection required"
                )
            
            if pressure > 8.5 or pressure < 3.5:
                EquipmentAlert.objects.create(
                    equipment_name=row['Equipment Name'],
                    alert_type='CRITICAL',
                    parameter='Pressure',
                    value=pressure,
                    threshold=8.5 if pressure > 8.5 else 3.5,
                    message=f"Pressure {'critically high' if pressure > 8.5 else 'critically low'}: {pressure:.2f} bar",
                    recommendation="Check pressure regulators immediately"
                )

        EquipmentRanking.objects.all().delete()
        params = EquipmentParameter.objects.filter(dataset=dataset).order_by('-health_score')
        for rank, param in enumerate(params, start=1):
            EquipmentRanking.objects.create(
                equipment_name=param.equipment_name,
                equipment_type=param.equipment_type,
                overall_score=param.health_score,
                efficiency_rank=rank,
                reliability_rank=rank,
                performance_rank=rank
            )

        return Response(summary)
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Upload error: {error_details}")
        return Response({
            "error": f"Failed to process file: {str(e)}"
        }, status=500)


@api_view(['GET'])
def get_trends(request):
    days = int(request.GET.get('days', 30))
    datasets = Dataset.objects.filter(
        uploaded_at__gte=datetime.now() - timedelta(days=days)
    ).order_by('uploaded_at')
    
    trends = {
        'dates': [],
        'flowrate': [],
        'pressure': [],
        'temperature': []
    }
    
    for ds in datasets:
        trends['dates'].append(ds.uploaded_at.strftime('%Y-%m-%d'))
        trends['flowrate'].append(float(ds.avg_flowrate))
        trends['pressure'].append(float(ds.avg_pressure))
        trends['temperature'].append(float(ds.avg_temperature))
    
    return Response(trends)


@api_view(['GET'])
def generate_pdf(request):
    dataset = Dataset.objects.order_by('-uploaded_at').first()

    if not dataset:
        return Response({"error": "No data available"}, status=400)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=1
    )
    
    story.append(Paragraph("Chemical Equipment Analysis Report", title_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Records', str(dataset.total_records)],
        ['Average Flowrate', f'{dataset.avg_flowrate:.2f} L/min'],
        ['Average Pressure', f'{dataset.avg_pressure:.2f} bar'],
        ['Average Temperature', f'{dataset.avg_temperature:.2f}°C'],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(summary_table)
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="equipment_report.pdf"'

    return response


@api_view(['GET'])
def export_to_excel(request):
    try:
        from .models import Dataset, EquipmentParameter, EquipmentRanking, EquipmentAlert
        
        dataset = Dataset.objects.order_by('-uploaded_at').first()
        
        if not dataset:
            return Response({'error': 'No data available'}, status=400)
        
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws_summary['A1'] = "Equipment Analysis Report"
        ws_summary['A1'].font = Font(bold=True, size=16, color="667EEA")
        ws_summary.merge_cells('A1:B1')
        
        ws_summary['A3'] = "Generated"
        ws_summary['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        ws_summary['A5'] = "Metric"
        ws_summary['B5'] = "Value"
        ws_summary['A5'].fill = header_fill
        ws_summary['B5'].fill = header_fill
        ws_summary['A5'].font = header_font
        ws_summary['B5'].font = header_font
        
        summary_data = [
            ["Total Records", dataset.total_records],
            ["Average Flowrate (L/min)", round(dataset.avg_flowrate, 2)],
            ["Average Pressure (bar)", round(dataset.avg_pressure, 2)],
            ["Average Temperature (°C)", round(dataset.avg_temperature, 2)],
        ]
        
        row = 6
        for metric, value in summary_data:
            ws_summary[f'A{row}'] = metric
            ws_summary[f'B{row}'] = value
            row += 1
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="equipment_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Excel export error: {error_details}")
        return Response({
            'error': f'Failed to generate Excel report: {str(e)}',
            'details': error_details
        }, status=500)


@api_view(['GET'])
def get_alerts(request):
    resolved = request.GET.get('resolved', 'false').lower() == 'true'
    alerts = EquipmentAlert.objects.filter(resolved=resolved).order_by('-created_at')[:50]
    
    data = []
    for alert in alerts:
        data.append({
            'id': alert.id,
            'equipment_name': alert.equipment_name,
            'alert_type': alert.alert_type,
            'parameter': alert.parameter,
            'value': alert.value,
            'threshold': alert.threshold,
            'message': alert.message,
            'recommendation': alert.recommendation,
            'created_at': alert.created_at,
            'resolved': alert.resolved,
            'resolved_at': alert.resolved_at,
            'predicted_failure_date': alert.predicted_failure_date,
            'confidence_score': alert.confidence_score
        })
    
    return Response(data)


@api_view(['POST'])
def resolve_alert(request, alert_id):
    try:
        alert = EquipmentAlert.objects.get(id=alert_id)
        alert.resolved = True
        alert.resolved_at = datetime.now()
        alert.save()
        return Response({'success': True, 'message': 'Alert resolved'})
    except EquipmentAlert.DoesNotExist:
        return Response({'error': 'Alert not found'}, status=404)


@api_view(['POST'])
def compare_equipment(request):
    equipment_names = request.data.get('equipment_names', [])
    
    if len(equipment_names) < 2:
        return Response({'error': 'Select at least 2 equipment'}, status=400)
    
    latest_dataset = Dataset.objects.order_by('-uploaded_at').first()
    
    if not latest_dataset:
        return Response({'error': 'No data available'}, status=400)
    
    equipment_data = EquipmentParameter.objects.filter(
        dataset=latest_dataset,
        equipment_name__in=equipment_names
    )
    
    comparison = []
    for eq in equipment_data:
        comparison.append({
            'name': eq.equipment_name,
            'type': eq.equipment_type,
            'flowrate': eq.flowrate,
            'pressure': eq.pressure,
            'temperature': eq.temperature,
            'health_score': eq.health_score
        })
    
    return Response(comparison)


@api_view(['GET'])
def get_maintenance_schedule(request):
    schedules = MaintenanceSchedule.objects.filter(
        status__in=['SCHEDULED', 'IN_PROGRESS']
    ).order_by('scheduled_date')[:20]
    
    data = []
    for schedule in schedules:
        data.append({
            'id': schedule.id,
            'equipment_name': schedule.equipment_name,
            'equipment_type': schedule.equipment_type,
            'scheduled_date': schedule.scheduled_date,
            'priority': schedule.priority,
            'status': schedule.status,
            'estimated_hours': schedule.estimated_hours,
            'parts_needed': schedule.parts_needed,
            'description': schedule.description
        })
    
    return Response(data)


@api_view(['POST'])
def create_maintenance_schedule(request):
    schedule = MaintenanceSchedule.objects.create(
        equipment_name=request.data.get('equipment_name'),
        equipment_type=request.data.get('equipment_type'),
        scheduled_date=request.data.get('scheduled_date'),
        priority=request.data.get('priority', 'MEDIUM'),
        estimated_hours=request.data.get('estimated_hours', 2),
        parts_needed=request.data.get('parts_needed', []),
        description=request.data.get('description', '')
    )
    
    return Response({'success': True, 'id': schedule.id})


@api_view(['POST'])
def update_maintenance_status(request, schedule_id):
    try:
        schedule = MaintenanceSchedule.objects.get(id=schedule_id)
        new_status = request.data.get('status')
        
        schedule.status = new_status
        if new_status == 'COMPLETED':
            schedule.completed_at = datetime.now()
        
        schedule.save()
        return Response({'success': True})
    except MaintenanceSchedule.DoesNotExist:
        return Response({'error': 'Schedule not found'}, status=404)


@api_view(['GET'])
def get_equipment_rankings(request):
    rankings = EquipmentRanking.objects.all()[:20]
    
    data = []
    for i, ranking in enumerate(rankings, start=1):
        data.append({
            'rank': i,
            'equipment_name': ranking.equipment_name,
            'equipment_type': ranking.equipment_type,
            'overall_score': ranking.overall_score,
            'efficiency_rank': ranking.efficiency_rank,
            'reliability_rank': ranking.reliability_rank,
            'performance_rank': ranking.performance_rank
        })
    
    return Response(data)


@api_view(['POST'])
def schedule_email_report(request):
    return Response({'success': True, 'message': 'Email report scheduled'})


@api_view(['GET'])
def get_email_schedules(request):
    return Response([])


@api_view(['POST'])
def update_email_schedule(request, schedule_id):
    return Response({'success': True})


@api_view(['DELETE'])
def delete_email_schedule(request, schedule_id):
    return Response({'success': True})